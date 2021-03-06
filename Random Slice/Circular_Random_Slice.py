# The final working prototype to fulfill my summer NSERC. This program reads in a series of CT images, finds a CIRCULAR
# dataset boundary, generates random cubes within the boundary sized by an estimated REV computed for the sample (based
# on total porosity) and slices the cubes at any angle between 0 and 360 degs. These angled slices through the cube are
# measured for porosity, the subsequent values get exported to an excel sheet.

import cv2
import glob
import numpy as np
import math
import random
import openpyxl
from openpyxl import load_workbook

est_por = None


# Main delegates the loading of image files (specified by a user), handles loading and/or creation of an excel sheet,
# controls the creation of metadata about the image files (i.e boundaries of data sets, REV of a cube, and dimensions
# of images) and finally handles the main loop of cube slicing and data processing.
def main():

	wb_ws_save = excel_handler()
	wb = wb_ws_save[0]
	ws = wb_ws_save[1]
	save_as = wb_ws_save[2]

	angles = [0, 15, 30, 45, 60, 75, 90, 105, 120, 135, 150, 165]  # Default angles
	# angles = [i for i in range(1, 180)]

	while True:
		try:
			number_of_cycles = int(raw_input("How many cycles do you wish to do: "))
		except ValueError:
			print("That was not a valid number....")
			continue
		break
	while True:
		try:
			custom_angles = (raw_input("Would you like to input custom angles Y or N: "))
			if custom_angles == "Y" or custom_angles == "y":
				angles = raw_input("What angles would you like to slice on: ")
				angles = map(int, angles.split())
		except ValueError:
			print("These are not valid angles")
			continue
		break

	images = file_reader()

	height = images[1].shape[0]
	width = images[1].shape[1]
	center = (int(height // 2), int(width // 2))
	radius = radius_finder(images, width, center)

	c_len = rev_finder(images, radius, center)

	# Ensures a odd cube size for the random slicer (Required for a correct random slice procedure to occur).
	if c_len % 2 == 0:
		c_len -= 1

	for i in range(0, number_of_cycles):
		vertex = vertex_generator(center, radius, c_len)
		data = cube_generator(vertex, images, c_len, len(images))
		cube_slicer(data[0], c_len, vertex, data[1], ws, angles)

	wb.save(save_as)


# Reads in image files specified by the users
def file_reader():
	while True:
		file_location = raw_input("Please specify the file path to where your images are stored: ")
		file_type = raw_input("Please specify the file type. I.e bmp: ")
		files = (glob.glob(file_location + "/*." + file_type))
		images = []
		for img_file in files:
			images.append(cv2.imread(img_file, 0))
		if len(images) == 0:
			print("The folder location or file type you selected were invalid, please try again")
			continue
		break

	return images


def excel_handler():
	valid_file = False

	while not valid_file:
		old_excel = raw_input("Would you like to use an existing excel file Y OR N: ")
		if old_excel == "Y" or old_excel == "y":
			excel_file = raw_input("Please specify the file name with proper suffix i.e sheet1.xlsx: ")
			try:
				wb = load_workbook(excel_file)
				ws = wb.create_sheet(wb.worksheets[0].title + " %s" % (len(wb.worksheets)))
				save_as = excel_file
				valid_file = True
			except IOError:
				print("Your specified file does not exist, please try again")

		elif old_excel == "N" or old_excel == "n":
			valid_file = True
			wb = openpyxl.Workbook()
			save_as = raw_input("What would you like to save the excel book as: ")
			save_as = save_as + ".xlsx"
			for chars in save_as:
				if chars in [':', '*', '?', '"', '<', '>', '|'] or ord(chars) == 92 or ord(chars) == 47:
					print("This is not a valid file name, it contains a :,*,?,<,>,|,/ etc. ")
					valid_file = False
			ws = wb.active

		else:
			print("Sorry, I didn't understand that.")
			continue

	return wb, ws, save_as


# Returns a radius such that everything within the circle is part of the imaged sample (i.e puts a upper limit on
# the boundary of our CT data) ensuring that we do not included any invalid pixels during future computations.
def radius_finder(images, width, center):
	radii = []

	for i in range(0, len(images), int(math.floor(len(images) // 10))):  # Take one measurement each tenth
		width_index = width
		cur_i = images[i]
		cur_p = cur_i[center[0], width - 1]  # Start on the right hand boundary

		while cur_p == 0 and width_index > 0:  # Continue iterating until we hit our data set, maxes out at left edge
			width_index -= 1
			cur_p = cur_i[center[0], width_index]
		radii.append(width_index - center[1])

	return max(radii)


# Determines whether a pixel (x,y co-ord) is within the range of a circle delimiting the boundary of our data set.
def is_in_circle(center, radius, vertices):
	for i in range(0, len(vertices)):
		if np.sqrt((vertices[i][0] - center[1]) ** 2 + (vertices[i][1] - center[0]) ** 2) > radius:
			return False
	return True


# Generates a vertex and checks to make sure all points of a cube emanating from said vertex fit inside our data set
# at which point a set of 4 vertices (the base of the cubic section) will be returned. This code will only get stuck
# if there are no cubes (within reason) with c_len = rev that fit in the given data set.
def vertex_generator(center, radius, c_len):
	# Edges are in reference to the perimeter of the circle delineating the data set boundary
	left_edge = center[1] - radius
	right_edge = center[1] + radius

	adders = [[0, 0], [0, c_len], [c_len, 0], [c_len, c_len]]

	vertices_in_circle = False
	loop_counter = 0

	while not vertices_in_circle:
		loop_counter += 1
		vertices = []
		bottom_left_coord = [random.randrange(left_edge, right_edge), random.randrange(left_edge, right_edge)]
		for i in range(4):
			vertices.append([a + b for a, b in zip(bottom_left_coord, adders[i])])
		vertices_in_circle = is_in_circle(center, radius, vertices)
		if loop_counter >= 50000:
			print(
				"This dataset is incompatible with the program, i.e there are no cubes of side length %d that will fit"
				" in the provided image data. Please try again with a larger and more regular image set if possible. "
				% c_len)
			exit()

	return vertices[0]


# Generates a cubic section within the dataset given a vertex
def cube_generator(vertex, images, c_len, stack_height):
	upper_3d_bound = stack_height - c_len
	z_position = random.randrange(0, upper_3d_bound)

	cube = np.zeros([c_len, c_len, c_len], int)
	x1 = vertex[0]
	y1 = vertex[1]
	z1 = z_position

	for z2 in range(z_position, z_position + c_len):
		for x2 in range(vertex[0], vertex[0] + c_len):
			for y2 in range(vertex[1], vertex[1] + c_len):
				cube[z2 - z1][x2 - x1][y2 - y1] = images[z2][y2][x2]

	data = [cube, z_position]

	return data


# This method takes in a odd sized cube and oversees the generation of various slice planes through it
def cube_slicer(cube, c_len, vertex, z_position, ws, angles):
	porosities = []

	slice_plane = np.zeros([c_len ** 2])
	mid_point = ((c_len ** 2) / 2) - (((c_len ** 2) / 2) // c_len)  # Middle of a 1D array of length n*n
	mid_indices = mid_point / c_len  # Middle of a 1D array of length n i.e an x,y or z array in 3D space

	for i in range(mid_point, mid_point + c_len):
		slice_plane[i] = cube[mid_indices][i - mid_point][mid_indices]  # Fill middle of plane with cubes values

	slopes = slope_generator(angles)

	for i in range(0, len(slopes)):
		slice_plane = slice_builder(cube, slice_plane, slopes[i], mid_point + c_len, mid_indices)
		porosities.append(por_calc(np.count_nonzero(slice_plane), float(len(slice_plane))))

	data_writer(ws, porosities, vertex, z_position, angles, cube, c_len)


# Iterates through a cube, appending rows/cols to a slice plane in order to simulate angular slicing
def slice_builder(cube, slice_plane, slope, slice_position, mid_indices):
	c_len = len(cube)
	z_position = mid_indices
	x_position = mid_indices
	max_slice = (c_len - 1) / 2

	loop_counter = 0
	decimal_tracker = 0

	# For angles 0 - 45 and 135 - 180
	if abs(slope) <= 1:
		# Build the top half of the slice plane
		while loop_counter < max_slice:
			if abs(slope) > 0 and slope % 1 == 0:  # Special Case for 45/135 degree angles
				decimal_tracker += 1
			else:
				decimal_tracker += math.modf(abs(slope))[0]
			if decimal_tracker >= 1:
				z_position += 1
				decimal_tracker -= 1
			x_position += 1
			for i in range(slice_position, slice_position + c_len):
				slice_plane[i] = cube[z_position][i - slice_position][
					x_position if slope >= 0 else c_len - x_position]  # If slope is negative switch side
			slice_position += c_len
			loop_counter += 1

		# Reset the loop position
		slice_position = (slice_position - (loop_counter + 1) * c_len)
		z_position = mid_indices
		x_position = mid_indices
		loop_counter = 0

		# Build the bottom half of the slice plane
		while loop_counter < max_slice:
			if abs(slope) > 0 and slope % 1 == 0:  # Ensures whole #'s get incremented in z
				decimal_tracker += 1
			else:
				decimal_tracker += math.modf(abs(slope))[0]
			if decimal_tracker >= 1:
				z_position -= 1
				decimal_tracker -= 1
			x_position -= 1
			for i in range(slice_position - c_len, slice_position):
				slice_plane[i] = cube[z_position][i - (slice_position - c_len)][
					x_position if slope >= 0 else c_len - (x_position + 1)]  # If slope is negative switch side
			slice_position -= c_len
			loop_counter += 1

	# For angles 46 - 134
	else:
		slope = 1/slope
		# Build the top half of the slice plane
		while loop_counter < max_slice:
			decimal_tracker += math.modf(abs(slope))[0]
			if decimal_tracker >= 1:
				x_position += 1
				decimal_tracker -= 1
			z_position += 1
			for i in range(slice_position, slice_position + c_len):
				slice_plane[i] = cube[z_position][i - slice_position][
					x_position if slope >= 0 else c_len - x_position]  # If slope is negative switch which side we do!
			slice_position += c_len
			loop_counter += 1

		# Reset the loop position
		slice_position = (slice_position - (loop_counter + 1) * c_len)
		z_position = mid_indices
		x_position = mid_indices
		loop_counter = 0

		# Build the bottom half of the slice plane
		while loop_counter < max_slice:
			decimal_tracker += math.modf(abs(slope))[0]
			if decimal_tracker >= 1:
				x_position -= 1
				decimal_tracker -= 1
			z_position -= 1
			for i in range(slice_position - c_len, slice_position):
				slice_plane[i] = cube[z_position][i - (slice_position - c_len)][
					x_position if slope >= 0 else c_len - (x_position + 1)]  # If slope is negative switch which side we do!
			slice_position -= c_len
			loop_counter += 1

	return slice_plane


# Given an array of angles, returns the corresponding slopes each angle produces
def slope_generator(angles):
	slopes = []

	for angle in angles:
		radian = math.radians(angle)
		slopes.append(round(math.tan(radian), 3))

	return slopes


# Writes the porosity data to an excel file
def data_writer(ws, porosities, vertex, z_position, angle, cube, c_len, counter=[0]):
	counter[0] += 1

	# If its the first time opening the sheet, write the angle information on the top row
	if counter[0] == 1:
		ws.cell(row=1, column=1).value = "Angle (x, y, z)"
		for i in range(1, len(angle) + 1):
			ws.cell(row=1, column=i + 1).value = angle[i - 1]
		ws.cell(row=1, column=i + 5).value = "c_len:"
		ws.cell(row=1, column=i + 6).value = c_len
		ws.cell(row=1, column=i + 2).value = "cube porosity"
		ws.cell(row=2, column=i + 5).value = "total porosity:"
		ws.cell(row=2, column=i + 6).value = est_por

	ws.cell(row=counter[0] + 1, column=1).value = "Slice at (%i,%i,%i)" % (vertex[0], vertex[1], z_position)
	for i in range(1, len(porosities) + 1):
		ws.cell(row=counter[0] + 1, column=i + 1).value = porosities[i - 1]
	ws.cell(row=counter[0] + 1, column=i + 2).value = por_calc(cube_porosity_counter(cube, c_len), c_len**3)


# Returns a approximate REV which I will use as the length of my cubes. Based on a line growing algorithm, which is
# rooted in the assumption that a 1D REV will translate in a 3D system, which will work for homogeneous samples only.
def rev_finder(images, radius, center):
	global est_por
	
	total_nonzero_pix = 0

	# Count total porosity for the data set
	for i in range(0, len(images)):
		total_nonzero_pix += np.count_nonzero(images[i])

	volume = math.pi * (radius ** 2) * i
	total_porosity = por_calc(total_nonzero_pix, volume)
	est_por = total_porosity

	# Grow a line until it contains a similar porosity to the total i.e a line of REV

	line_holder = []

	for j in range(0, 1000):
		random_image = random.randrange(0, len(images))

		line = [images[random_image][center[0]][center[1]]]
		gi = 0  # Growth Incrementer

		while por_calc(np.count_nonzero(line), len(line)) < total_porosity - 1 and gi < center[1] - 1 or \
			por_calc(np.count_nonzero(line), len(line)) > total_porosity + 1 and gi < center[1] - 1:
			gi += 1
			line.extend([images[random_image][center[0]][center[1] - gi]])
			line.extend([images[random_image][center[0]][center[1] + gi]])

		line_holder.append(len(line))

	return sum(line_holder) / len(line_holder)


# Given total bright and dark pixels of an image(s), returns the porosity as a percentage
def por_calc(bright_pixels, total_pixels):
	return (1 - (bright_pixels / float(total_pixels))) * 100


# A simple function used to count the amount of pore space present in a cube
def cube_porosity_counter(cube, c_len):
	count = 0

	for i in range(c_len):
		for j in range(c_len):
			for k in range(c_len):
				count += np.count_nonzero(cube[i][j][k])

	return count


main()
